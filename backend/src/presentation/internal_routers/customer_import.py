"""
Customer Import API Endpoint

This endpoint handles bulk import of customers from CSV/Excel files.
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from typing import List, Dict, Any
import csv
import io
import uuid
from datetime import datetime
import logging
from openpyxl import load_workbook

from ...config.database import get_db
from ...presentation.dependencies.auth import get_current_user, get_tenant_context
from ...config.database import User
from ...config.crm_models import Customer
from ...config.crm_crud import create_customer, get_customer_by_email
from ...models.crm import CustomerCreate

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/crm", tags=["Customer Import"])

class CustomerImportResponse:
    def __init__(self, success: bool, message: str, imported_count: int = 0, 
                 failed_count: int = 0, errors: List[str] = None):
        self.success = success
        self.message = message
        self.imported_count = imported_count
        self.failed_count = failed_count
        self.errors = errors or []

@router.post("/customers/import")
async def import_customers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    tenant_context: dict = Depends(get_tenant_context)
):
    """
    Import customers from CSV or Excel file
    
    Expected columns:
    - firstName, lastName, email (required)
    - phone, mobile, cnic, dateOfBirth, gender
    - address, city, state, country, postalCode
    - customerType, customerStatus, creditLimit, currentBalance
    - paymentTerms, notes, tags
    """
    try:
        if not tenant_context:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Tenant context required"
            )
        
        tenant_id = tenant_context["tenant_id"]
        
        # Validate file type
        if not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No file provided"
            )
        
        file_extension = file.filename.lower().split('.')[-1]
        if file_extension not in ['csv', 'xlsx', 'xls']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="File must be CSV or Excel format"
            )
        
        # Read file content
        file_content = await file.read()
        
        # Parse file based on extension
        try:
            if file_extension == 'csv':
                csv_reader = csv.DictReader(io.StringIO(file_content.decode('utf-8')))
                rows = list(csv_reader)
            else:  # Excel files
                workbook = load_workbook(io.BytesIO(file_content))
                worksheet = workbook.active
                
                # Get headers from first row
                headers = [cell.value for cell in worksheet[1]]
                
                # Convert to list of dictionaries
                rows = []
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        row_dict = {}
                        for i, value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = value
                        rows.append(row_dict)
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Error parsing file: {str(e)}"
            )
        
        if not rows:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No data found in file"
            )
        
        # Validate required columns
        required_columns = ['firstName', 'lastName', 'email']
        available_columns = list(rows[0].keys()) if rows else []
        missing_columns = [col for col in required_columns if col not in available_columns]
        if missing_columns:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        # Process customers
        imported_count = 0
        failed_count = 0
        errors = []
        
        for index, row in enumerate(rows):
            try:
                # Helper function to safely get string values
                def safe_str(value, default=''):
                    if value is None or value == '':
                        return default
                    return str(value).strip()
                
                # Helper function to safely get numeric values
                def safe_float(value, default=0.0):
                    if value is None or value == '':
                        return default
                    try:
                        return float(value)
                    except (ValueError, TypeError):
                        return default
                
                # Prepare customer data
                customer_data = {
                    "firstName": safe_str(row.get('firstName')),
                    "lastName": safe_str(row.get('lastName')),
                    "email": safe_str(row.get('email')).lower(),
                    "phone": safe_str(row.get('phone')) if row.get('phone') else None,
                    "mobile": safe_str(row.get('mobile')) if row.get('mobile') else None,
                    "cnic": safe_str(row.get('cnic')) if row.get('cnic') else None,
                    "dateOfBirth": row.get('dateOfBirth') if row.get('dateOfBirth') else None,
                    "gender": safe_str(row.get('gender')).lower() if row.get('gender') else None,
                    "address": safe_str(row.get('address')) if row.get('address') else None,
                    "city": safe_str(row.get('city')) if row.get('city') else None,
                    "state": safe_str(row.get('state')) if row.get('state') else None,
                    "country": safe_str(row.get('country'), 'Pakistan'),
                    "postalCode": safe_str(row.get('postalCode')) if row.get('postalCode') else None,
                    "customerType": safe_str(row.get('customerType'), 'individual').lower(),
                    "customerStatus": safe_str(row.get('customerStatus'), 'active').lower(),
                    "creditLimit": safe_float(row.get('creditLimit')),
                    "currentBalance": safe_float(row.get('currentBalance')),
                    "paymentTerms": safe_str(row.get('paymentTerms'), 'Cash'),
                    "notes": safe_str(row.get('notes')) if row.get('notes') else None,
                    "tags": []
                }
                
                # Validate required fields
                if not customer_data["firstName"] or not customer_data["lastName"] or not customer_data["email"]:
                    errors.append(f"Row {index + 2}: Missing required fields (firstName, lastName, email)")
                    failed_count += 1
                    continue
                
                # Validate email format
                if '@' not in customer_data["email"]:
                    errors.append(f"Row {index + 2}: Invalid email format")
                    failed_count += 1
                    continue
                
                # Check if customer already exists
                existing_customer = get_customer_by_email(customer_data["email"], db, tenant_id)
                if existing_customer:
                    errors.append(f"Row {index + 2}: Customer with email {customer_data['email']} already exists")
                    failed_count += 1
                    continue
                
                # Validate customer type
                if customer_data["customerType"] not in ['individual', 'business']:
                    customer_data["customerType"] = 'individual'
                
                # Validate customer status
                if customer_data["customerStatus"] not in ['active', 'inactive', 'blocked']:
                    customer_data["customerStatus"] = 'active'
                
                # Validate gender
                if customer_data["gender"] and customer_data["gender"] not in ['male', 'female', 'other']:
                    customer_data["gender"] = None
                
                # Process tags if provided
                if row.get('tags'):
                    tags_str = safe_str(row.get('tags'))
                    if tags_str:
                        customer_data["tags"] = [tag.strip() for tag in tags_str.split(',') if tag.strip()]
                
                # Generate customer ID
                customer_data["customerId"] = f"CUST{str(uuid.uuid4())[:8].upper()}"
                
                # Create customer
                db_customer = create_customer(db, customer_data, tenant_id)
                if db_customer:
                    imported_count += 1
                    logger.info(f"Imported customer: {customer_data['firstName']} {customer_data['lastName']} ({customer_data['email']})")
                else:
                    errors.append(f"Row {index + 2}: Failed to create customer")
                    failed_count += 1
                    
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                failed_count += 1
                logger.error(f"Error processing row {index + 2}: {str(e)}")
        
        # Prepare response
        if imported_count > 0:
            message = f"Successfully imported {imported_count} customers"
            if failed_count > 0:
                message += f", {failed_count} failed"
        else:
            message = "No customers were imported"
        
        return CustomerImportResponse(
            success=imported_count > 0,
            message=message,
            imported_count=imported_count,
            failed_count=failed_count,
            errors=errors[:10]  # Limit errors to first 10
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error in customer import: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Import failed: {str(e)}"
        )

@router.get("/customers/import/template")
async def download_import_template():
    """
    Download a template CSV file for customer import
    """
    try:
        # Create template data
        template_data = [
            {
                'firstName': 'John',
                'lastName': 'Doe',
                'email': 'john.doe@example.com',
                'phone': '+92-300-1234567',
                'mobile': '+92-300-1234567',
                'cnic': '12345-1234567-1',
                'dateOfBirth': '1990-01-01',
                'gender': 'male',
                'address': '123 Main Street',
                'city': 'Karachi',
                'state': 'Sindh',
                'country': 'Pakistan',
                'postalCode': '75000',
                'customerType': 'individual',
                'customerStatus': 'active',
                'creditLimit': '10000',
                'currentBalance': '0',
                'paymentTerms': 'Cash',
                'notes': 'VIP Customer',
                'tags': 'premium,vip'
            },
            {
                'firstName': 'Jane',
                'lastName': 'Smith',
                'email': 'jane.smith@example.com',
                'phone': '+92-300-7654321',
                'mobile': '+92-300-7654321',
                'cnic': '12345-7654321-2',
                'dateOfBirth': '1985-05-15',
                'gender': 'female',
                'address': '456 Oak Avenue',
                'city': 'Lahore',
                'state': 'Punjab',
                'country': 'Pakistan',
                'postalCode': '54000',
                'customerType': 'business',
                'customerStatus': 'active',
                'creditLimit': '25000',
                'currentBalance': '0',
                'paymentTerms': 'Credit',
                'notes': 'Regular Customer',
                'tags': 'regular,standard'
            }
        ]
        
        # Convert to CSV using built-in csv module
        csv_buffer = io.StringIO()
        if template_data:
            fieldnames = template_data[0].keys()
            writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(template_data)
        
        csv_content = csv_buffer.getvalue()
        
        return {
            "success": True,
            "message": "Template generated successfully",
            "template": csv_content,
            "filename": "customer_import_template.csv"
        }
        
    except Exception as e:
        logger.error(f"Error generating template: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate template: {str(e)}"
        )
